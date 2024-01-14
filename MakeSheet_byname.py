import openpyxl
"""
1枚目のワークシートをwsとします
wsはセルA1にカラム名
セルA2から下へと氏名が並んで行くようなテンプレートとします
"""
filepath = "sample.xlsx" #ファイル名に応じて適宜変更してください

def mkws_byname(filepath):
  wb = openpyxl.load_workbook(filepath)
  ws = wb.worksheets[0] #1枚目のsheetをwsとする

    #重複しないようにシートを作成
  namelist = []
  for i in range(2, ws.max_row+1):
    name = ws.cell(i, 1).value #氏名の値を取ってくる
    if name not in namelist: #リストにない氏名のみリストに追加する
      namelist.append(name)

    #最後にnamelistの中の氏名でシートを作る
  for _ in namelist:
      wb.create_sheet(title=_)
  wb.save("sample_mksheet.xlsx") #上書き保存

